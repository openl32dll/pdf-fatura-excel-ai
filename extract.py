#!/usr/bin/env python3
"""
PDF Invoice -> Excel + AI-Powered Categorization Tool
======================================================

Reads a PDF invoice, extracts its line items (description/qty/unit price/
total) as a table, categorizes each item, and produces a formatted Excel
report (detail + summary sheet).

Two modes:
  - AI mode:      if ANTHROPIC_API_KEY is set, categorization is done via
                   Claude (much more accurate on vague/varied item
                   descriptions).
  - Fallback mode: without an API key, falls back to keyword matching —
                   works end-to-end with no API key required.

Output language: --lang en (default) or --lang tr — controls the column
headers and category names in the generated Excel file, independent of
the language the source invoice is written in.

Usage:
    python extract.py sample_data/invoice_sample_en.pdf -o output/report.xlsx
    python extract.py sample_data/fatura_ornek.pdf -o output/rapor.xlsx --lang tr
    python extract.py invoice.pdf -o report.xlsx --no-ai   # force AI off
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# Language packs — column headers, category names, keyword rules, labels
# --------------------------------------------------------------------------

LANGS = {
    "en": {
        "categories": ["Office Supplies", "Software/Licenses", "Transport", "Meals/Catering", "Consulting", "Other"],
        "keyword_map": {
            "Software/Licenses": ["software", "license", "cloud", "subscription", "saas"],
            "Transport": ["taxi", "uber", "flight", "fuel", "gas", "transport", "transfer", "mileage"],
            "Meals/Catering": ["meal", "coffee", "catering", "restaurant", "lunch", "dinner"],
            "Consulting": ["consult", "consulting", "service fee", "advisory"],
            "Office Supplies": ["paper", "pen", "mouse", "keyboard", "office", "toner", "folder", "stationery"],
        },
        "columns": ["Description", "Qty", "Unit Price", "Total", "Category"],
        "sheet_detail": "Invoice Detail",
        "sheet_summary": "Summary",
        "summary_labels": {"invoice_no": "Invoice No", "date": "Date", "vendor": "Vendor"},
        "summary_headers": ["Category", "Total"],
        "grand_total": "GRAND TOTAL",
        "currency_format": "#,##0.00",
        "header_patterns": {
            "invoice_no": r"Invoice\s*(?:No\.?|Number):?\s*(.+)",
            "date": r"Date:?\s*([\d./-]+)",
            "vendor": r"Vendor:?\s*(.+)",
        },
        "table_header_markers": ["description"],
        "other_category": "Other",
    },
    "tr": {
        "categories": ["Ofis Malzemesi", "Yazılım/Lisans", "Ulaşım", "Yemek/İkram", "Danışmanlık", "Diğer"],
        "keyword_map": {
            "Yazılım/Lisans": ["yazılım", "lisans", "bulut", "abonelik", "software", "subscription"],
            "Ulaşım": ["taksi", "otobüs", "uçak", "yakıt", "benzin", "ulaşım", "transfer"],
            "Yemek/İkram": ["yemek", "kahve", "ikram", "catering", "restoran"],
            "Danışmanlık": ["danışman", "consulting", "hizmet bedeli"],
            "Ofis Malzemesi": ["kağıt", "kalem", "mouse", "klavye", "ofis", "toner", "dosya"],
        },
        "columns": ["Açıklama", "Miktar", "Birim Fiyat", "Tutar", "Kategori"],
        "sheet_detail": "Fatura Detay",
        "sheet_summary": "Özet",
        "summary_labels": {"invoice_no": "Fatura No", "date": "Tarih", "vendor": "Satıcı"},
        "summary_headers": ["Kategori", "Toplam Tutar"],
        "grand_total": "GENEL TOPLAM",
        "currency_format": "#,##0.00 TL",
        "header_patterns": {
            "invoice_no": r"Fatura No:?\s*(.+)",
            "date": r"Tarih:?\s*([\d./-]+)",
            "vendor": r"Sat[ıi]c[ıi]:?\s*(.+)",
        },
        "table_header_markers": ["açıklama", "aciklama"],
        "other_category": "Diğer",
    },
}


# --------------------------------------------------------------------------
# Raw extraction from the PDF
# --------------------------------------------------------------------------

def extract_text(pdf_path: Path) -> str:
    with pdfplumber.open(pdf_path) as pdf:
        return "\n".join(page.extract_text() or "" for page in pdf.pages)


def extract_line_items(pdf_path: Path, lang: dict) -> list[dict]:
    items = []
    markers = lang["table_header_markers"]
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                if not table:
                    continue
                header = [c.strip().lower() if c else "" for c in table[0]]
                if not any(marker in h for h in header for marker in markers):
                    continue
                for row in table[1:]:
                    if not row or len(row) < 4 or not row[0]:
                        continue
                    desc, qty, unit_price, total = row[:4]
                    items.append(
                        {
                            "description": desc.strip(),
                            "qty": _to_number(qty),
                            "unit_price": _to_number(unit_price),
                            "total": _to_number(total),
                        }
                    )
    return items


def extract_header_fields(text: str, lang: dict) -> dict:
    def find(pattern, default=""):
        m = re.search(pattern, text, flags=re.IGNORECASE)
        return m.group(1).strip() if m else default

    p = lang["header_patterns"]
    return {
        "invoice_no": find(p["invoice_no"]),
        "date": find(p["date"]),
        "vendor": find(p["vendor"]),
    }


def _to_number(raw: str) -> float:
    """Parses a number whether it's formatted EU-style (1.200,00) or
    US/UK-style (1,200.00) — auto-detects based on separator order/shape,
    so it works correctly regardless of the invoice's source language."""
    if raw is None:
        return 0.0
    cleaned = re.sub(r"[^\d.,-]", "", raw.strip())
    if not cleaned:
        return 0.0

    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")  # 1.200,00
        else:
            cleaned = cleaned.replace(",", "")  # 1,200.00
    elif "," in cleaned:
        # ambiguous: 2 digits after last comma -> decimal sep, else thousands sep
        tail = cleaned.split(",")[-1]
        cleaned = cleaned.replace(",", ".") if len(tail) == 2 else cleaned.replace(",", "")

    try:
        return float(cleaned)
    except ValueError:
        return 0.0


# --------------------------------------------------------------------------
# Categorization
# --------------------------------------------------------------------------

def categorize_with_keywords(items: list[dict], lang: dict) -> list[dict]:
    for item in items:
        desc_lower = item["description"].lower()
        category = lang["other_category"]
        for cat, keywords in lang["keyword_map"].items():
            if any(kw in desc_lower for kw in keywords):
                category = cat
                break
        item["category"] = category
    return items


def categorize_with_ai(items: list[dict], lang: dict) -> list[dict]:
    import anthropic

    client = anthropic.Anthropic()
    descriptions = [item["description"] for item in items]
    categories = lang["categories"]

    prompt = f"""Assign each of the following invoice line items to exactly
one of these categories: {", ".join(categories)}.

Items (in order):
{json.dumps(descriptions, ensure_ascii=False, indent=2)}

Return ONLY a JSON list in this exact format, nothing else:
["Category1", "Category2", ...]
The list length must exactly match the number of items, preserving order."""

    resp = client.messages.create(
        model="claude-sonnet-4-5",
        max_tokens=1000,
        messages=[{"role": "user", "content": prompt}],
    )
    raw = resp.content[0].text.strip()
    raw = re.sub(r"^```(json)?|```$", "", raw, flags=re.MULTILINE).strip()
    ai_categories = json.loads(raw)

    for item, cat in zip(items, ai_categories):
        item["category"] = cat if cat in categories else lang["other_category"]
    return items


# --------------------------------------------------------------------------
# Excel report
# --------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def build_excel(header: dict, items: list[dict], out_path: Path, lang: dict) -> None:
    wb = Workbook()

    # --- Detail sheet ---
    ws = wb.active
    ws.title = lang["sheet_detail"]
    cols = lang["columns"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for item in items:
        ws.append([item["description"], item["qty"], item["unit_price"], item["total"], item["category"]])

    for c in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 24
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=4):
        for cell in row:
            cell.number_format = lang["currency_format"]

    # --- Summary sheet ---
    ws2 = wb.create_sheet(lang["sheet_summary"])
    labels = lang["summary_labels"]
    ws2.append([labels["invoice_no"], header.get("invoice_no", "")])
    ws2.append([labels["date"], header.get("date", "")])
    ws2.append([labels["vendor"], header.get("vendor", "")])
    ws2.append([])
    ws2.append(lang["summary_headers"])
    for c in range(1, 3):
        cell = ws2.cell(row=5, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    totals: dict[str, float] = {}
    for item in items:
        totals[item["category"]] = totals.get(item["category"], 0.0) + item["total"]

    for cat, total in sorted(totals.items(), key=lambda kv: -kv[1]):
        ws2.append([cat, total])
    ws2.append([])
    ws2.append([lang["grand_total"], sum(item["total"] for item in items)])
    ws2["A" + str(ws2.max_row)].font = Font(bold=True)
    ws2["B" + str(ws2.max_row)].font = Font(bold=True)

    for row in ws2.iter_rows(min_row=6, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = lang["currency_format"]
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 18

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# --------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Converts a PDF invoice into an Excel report.")
    parser.add_argument("pdf", type=Path, help="Input PDF invoice file")
    parser.add_argument("-o", "--output", type=Path, default=Path("output/report.xlsx"))
    parser.add_argument("--lang", choices=["en", "tr"], default="en", help="Output language (column headers, categories)")
    parser.add_argument("--no-ai", action="store_true", help="Disable AI categorization")
    args = parser.parse_args()

    lang = LANGS[args.lang]

    if not args.pdf.exists():
        sys.exit(f"Error: {args.pdf} not found")

    text = extract_text(args.pdf)
    header = extract_header_fields(text, lang)
    items = extract_line_items(args.pdf, lang)

    if not items:
        sys.exit("Error: no recognizable invoice table found in the PDF.")

    use_ai = not args.no_ai
    if use_ai:
        try:
            items = categorize_with_ai(items, lang)
            mode = "AI (Claude)"
        except Exception as e:  # no API key / error -> fallback
            print(f"[warning] AI categorization unavailable ({e}); falling back to keyword mode.")
            items = categorize_with_keywords(items, lang)
            mode = "Keyword (fallback)"
    else:
        items = categorize_with_keywords(items, lang)
        mode = "Keyword (--no-ai)"

    build_excel(header, items, args.output, lang)

    print(f"✅ Report created: {args.output}")
    print(f"   Categorization mode: {mode}")
    print(f"   {len(items)} items processed, grand total: {sum(i['total'] for i in items):,.2f}")


if __name__ == "__main__":
    main()
